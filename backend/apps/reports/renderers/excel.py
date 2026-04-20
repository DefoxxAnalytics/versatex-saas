"""
Excel Renderer for reports.
Uses openpyxl for Excel generation with styled worksheets.
"""
from io import BytesIO
from .base import BaseRenderer

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Border, Side,
        Alignment, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelRenderer(BaseRenderer):
    """
    Renders reports as Excel workbooks with styled worksheets.
    Creates multi-sheet workbooks with formatted tables and charts.
    """

    # Brand colors
    HEADER_BG = '1E3A5F'  # Navy blue
    HEADER_TEXT = 'FFFFFF'  # White
    ACCENT_COLOR = '2563EB'  # Blue accent
    ALT_ROW_BG = 'F5F7FA'  # Light gray

    @property
    def content_type(self) -> str:
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @property
    def file_extension(self) -> str:
        return '.xlsx'

    def render(self) -> BytesIO:
        """Render report data as Excel workbook."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel rendering. "
                "Install it with: pip install openpyxl"
            )

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create styles
        header_style = self._create_header_style()
        data_style = self._create_data_style()

        # Add Summary sheet first
        self._add_summary_sheet(wb, header_style)

        # Add data sheets based on report content
        self._add_data_sheets(wb, header_style)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_header_style(self):
        """Create header cell style."""
        return {
            'font': Font(bold=True, color=self.HEADER_TEXT, size=11),
            'fill': PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _create_data_style(self):
        """Create data cell style."""
        return {
            'font': Font(size=10),
            'alignment': Alignment(vertical='center'),
            'border': Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
        }

    def _apply_style(self, cell, style_dict):
        """Apply style dictionary to a cell."""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)

    def _add_summary_sheet(self, wb, header_style):
        """Add summary/overview sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = self.metadata.get('report_title', self.report_name)
        ws['A1'].font = Font(bold=True, size=16, color=self.HEADER_BG)
        ws.merge_cells('A1:D1')

        # Metadata
        ws['A3'] = 'Organization:'
        ws['B3'] = self.metadata.get('organization', 'N/A')
        ws['A4'] = 'Period:'
        ws['B4'] = f"{self.metadata.get('period_start', 'N/A')} to {self.metadata.get('period_end', 'N/A')}"
        ws['A5'] = 'Generated:'
        ws['B5'] = self.metadata.get('generated_at', 'N/A')
        ws['A6'] = 'Report Type:'
        ws['B6'] = self.metadata.get('report_type', 'N/A')

        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)

        # Overview data
        overview = self.report_data.get('overview', {})
        if not overview:
            overview = self.report_data.get('summary', {})
            if not overview:
                overview = self.report_data.get('kpis', {})

        if overview:
            row = 8
            ws[f'A{row}'] = 'Key Metrics'
            ws[f'A{row}'].font = Font(bold=True, size=12, color=self.HEADER_BG)
            row += 1

            # Headers
            ws[f'A{row}'] = 'Metric'
            ws[f'B{row}'] = 'Value'
            self._apply_style(ws[f'A{row}'], header_style)
            self._apply_style(ws[f'B{row}'], header_style)
            row += 1

            for key, value in overview.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                if 'spend' in key.lower() or 'amount' in key.lower() or 'savings' in key.lower():
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].number_format = '$#,##0.00'
                elif 'percentage' in key.lower() or 'rate' in key.lower():
                    ws[f'B{row}'] = value / 100 if isinstance(value, (int, float)) else value
                    ws[f'B{row}'].number_format = '0.0%'
                elif isinstance(value, (int, float)):
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].number_format = '#,##0'
                else:
                    ws[f'B{row}'] = str(value)
                row += 1

        # Auto-fit columns
        self._auto_fit_columns(ws)

    def _add_data_sheets(self, wb, header_style):
        """Add data sheets for various report sections."""
        # Map of data keys to sheet names (max 31 chars for Excel)
        data_sections = {
            'spend_by_category': 'Spend by Category',
            'spend_by_supplier': 'Spend by Supplier',
            'top_suppliers': 'Top Suppliers',
            'top_categories': 'Top Categories',
            'monthly_trend': 'Monthly Trend',
            'pareto_data': 'Pareto Analysis',
            'tail_suppliers': 'Tail Suppliers',
            'consolidation_opportunities': 'Consolidation Opps',
            'category_opportunities': 'Category Opps',
            'stratification': 'Stratification',
            'compliance_summary': 'Compliance Summary',
            'violations': 'Violations',
            'savings_by_type': 'Savings by Type',
            'action_plan': 'Action Plan',
            'categories': 'Categories',
            'suppliers': 'Suppliers',
        }

        for key, sheet_name in data_sections.items():
            data = self.report_data.get(key, [])
            if data and isinstance(data, list) and len(data) > 0:
                self._create_data_sheet(wb, sheet_name, data, header_style)

    def _create_data_sheet(self, wb, sheet_name, data, header_style):
        """Create a worksheet from a list of dictionaries."""
        if not data or not isinstance(data[0], dict):
            return

        ws = wb.create_sheet(sheet_name[:31])  # Excel limit

        # Get headers from first item
        headers = list(data[0].keys())

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            self._apply_style(cell, header_style)

        # Write data
        for row_idx, item in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply number format based on header
                if 'amount' in header.lower() or 'spend' in header.lower() or 'savings' in header.lower():
                    cell.number_format = '$#,##0.00'
                elif 'percentage' in header.lower() or 'rate' in header.lower():
                    if isinstance(value, (int, float)) and value > 1:
                        cell.value = value / 100
                    cell.number_format = '0.0%'
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    if isinstance(value, float):
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '#,##0'

                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=self.ALT_ROW_BG,
                        end_color=self.ALT_ROW_BG,
                        fill_type='solid'
                    )

        # Auto-fit columns
        self._auto_fit_columns(ws)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add chart for certain data types
        if len(data) >= 3 and len(data) <= 20:
            if 'amount' in headers or 'spend' in headers:
                self._add_chart(ws, headers, len(data))

    def _auto_fit_columns(self, ws):
        """Auto-fit column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = None
            for cell in column_cells:
                try:
                    # Get column letter - handle merged cells
                    if column is None and hasattr(cell, 'column_letter'):
                        column = cell.column_letter
                    elif column is None and hasattr(cell, 'column'):
                        column = get_column_letter(cell.column)

                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            # Set width with some padding, max 50
            if column:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

    def _add_chart(self, ws, headers, data_count):
        """Add a bar chart to the worksheet."""
        # Find the value column
        value_col = None
        label_col = 1

        for i, header in enumerate(headers, 1):
            if header in ['amount', 'spend', 'total_spend', 'estimated_savings']:
                value_col = i
                break

        if not value_col:
            return

        chart = BarChart()
        chart.type = "col"
        chart.title = ws.title
        chart.style = 10

        # Data reference
        data_ref = Reference(ws, min_col=value_col, min_row=1, max_row=data_count + 1)
        cats_ref = Reference(ws, min_col=label_col, min_row=2, max_row=data_count + 1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4

        # Position chart
        chart_col = len(headers) + 2
        ws.add_chart(chart, f"{get_column_letter(chart_col)}2")
